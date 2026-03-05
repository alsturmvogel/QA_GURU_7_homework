import os
import csv
import zipfile
from zipfile import ZipFile
from io import TextIOWrapper
from pypdf import PdfReader
from openpyxl import load_workbook

# Пути к файлам
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RESOURCES_DIR = os.path.join(BASE_DIR, 'resources') # Склеивает BASE_DIR + "resources"
ARCHIVE_PATH = os.path.join(RESOURCES_DIR, 'archive.zip') #Склеивает RESOURCES_DIR + "archive.zip"


# Создаем архив
def create_archive():
    if not os.path.exists(RESOURCES_DIR):
        os.mkdir(RESOURCES_DIR)

    with ZipFile(ARCHIVE_PATH, 'w') as zip_file:
        zip_file.write(os.path.join(BASE_DIR, 'doc pdf.pdf'), arcname='doc pdf.pdf')
        zip_file.write(os.path.join(BASE_DIR, 'doc xlsx.xlsx'), arcname='doc xlsx.xlsx')
        zip_file.write(os.path.join(BASE_DIR, 'doc csv.csv'), arcname='doc csv.csv')

    print('Архив создан')

# Проверка PDF из архива
def test_pdf_in_archive():
    with zipfile.ZipFile(ARCHIVE_PATH) as zip_file:
        with zip_file.open('doc pdf.pdf') as pdf_file:
            reader = PdfReader(pdf_file)

            first_page = reader.pages[0]
            text = first_page.extract_text()

            assert 'Тестовый PDF файл' in text
            print('PDF проверка пройдена')

# Проверка XLSX из архива
def test_xlsx_in_archive():
    with zipfile.ZipFile(ARCHIVE_PATH) as zip_file:
        with zip_file.open('doc xlsx.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

            assert sheet.cell(row=16, column=3).value == 1  # проверяем ячейку C16
            print('XLSX проверка пройдена')

# Проверка CSV из архива
def test_csv_in_archive():
    with zipfile.ZipFile(ARCHIVE_PATH) as zip_file:
        with zip_file.open('doc csv.csv') as csv_file:
            csvreader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8-sig'), delimiter=';'))
            third_row = csvreader[2]

            assert third_row[1] == 'Отлично!'  # столбец B = индекс 1
            print('CSV проверка пройдена')

create_archive()
test_pdf_in_archive()
test_xlsx_in_archive()
test_csv_in_archive()